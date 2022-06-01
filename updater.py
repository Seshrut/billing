from openpyxl import load_workbook


def clear():
    print("\n"*30)


# load workbook
wb = load_workbook("DATABASE FOR BILLING.xlsx")
WB = load_workbook("DATABASE FOR MESSAGING.xlsx")
# load worksheet
bill = wb.active  # billing
message = wb.active  # messaging
# variables
max_bill = bill.max_row  # billing
max_mess = message.max_row  # messaging

for row in range(1, max_bill-2):
    name = bill["C"+str(row+3)].value
    message["A"+str(row+1)] = name
    B_day = bill["R"+str(row+3)].value
    message["B"+str(row+1)] = B_day
    B_mon = bill["S"+str(row+3)].value
    message["C"+str(row+1)] = B_mon
    ph_no = bill["B"+str(row+3)].value
    message["E"+str(row+1)] = ph_no
    APPEND = [name, B_day, B_mon, "B", ph_no, "U"]
    message.append(APPEND)
    message["D"+str(row+1)] = "B"
    message["F"+str(row+1)] = "U"
    WB.save("test.xlsx")
for row in range(max_bill-2, (max_bill-2)*2):
    name = bill["C"+str(row+3)].value
    message["A"+str(row+1)] = name
    A_day = bill["T"+str(row+3)].value
    message["B"+str(row+1)] = A_day
    A_mon = bill["U"+str(row+3)].value
    message["C"+str(row+1)] = A_mon
    ph_no = bill["B"+str(row+3)].value
    message["E"+str(row+1)] = ph_no
    APPEND = [name, A_day, A_mon, "A", ph_no, "U"]
    message.append(APPEND)
    message["D"+str(row+1)] = "A"
    message["F"+str(row+1)] = "U"
    WB.save("test.xlsx")
clear()