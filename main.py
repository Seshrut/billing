import os
try:
    import pandas as pd
except ModuleNotFoundError:
    os.system('pip install pandas')
try:
    import openpyxl
    from openpyxl import *
except ModuleNotFoundError:
    os.system('pip install openpyxl')
    os.system('pip install pywhatkit')
import datetime
from datetime import *
import time

def clear(lines=45):
    print("\n"*lines)


# Try except to make workbook or not
try:
    wb = load_workbook("DATABASE.xlsx")
    cust = wb["Customer"]
    bill = wb["Billing"]
    invoice = wb["Invoice"]
    art = wb["Articles"]
except Exception as new:
    print("Making a new DATABASE file \n Fill the Articles tab \n ARTICLES SHALL BE IN ASSENDING ORDER")
    wb = openpyxl.Workbook()
    cust = wb.active
    cust.title = "Customer"
    bill = wb.create_sheet(title="Billing")
    invoice = wb.create_sheet(title="Invoice")
    art = wb.create_sheet(title="Articles")

    cust["A1"].value = "Ph_No"
    cust["B1"].value = "Name"
    cust["C1"].value = "B Day"
    cust["D1"].value = "B Mon"
    cust["E1"].value = "U or S"
    cust["F1"].value = "A Day"
    cust["G1"].value = "A Mon"
    cust["H1"].value = "U or S"
    cust["I1"].value = "Frequency"
    cust["J1"].value = "PROFIT"

    bill["A1"].value = "Ph No"
    bill["B1"].value = "Date"
    bill["C1"].value = "Month"
    bill["D1"].value = "Hour"
    bill["E1"].value = "Minute"
    bill["F1"].value = "Final price"
    bill["G1"].value = "MOP"
    bill["H1"].value = "CP"
    bill["I1"].value = "Profit"
    bill["J1"].value = "Invoice"

    invoice["A1"].value = "Ph No"
    invoice["B1"].value = "Art No"
    invoice["C1"].value = "Art Desc"
    invoice["D1"].value = "QTY"
    invoice["E1"].value = "OUM"
    invoice["F1"].value = "MRP"
    invoice["G1"].value = "AMT"
    invoice["H1"].value = "Discount"
    invoice["I1"].value = "SP"
    invoice["J1"].value = "CP"
    invoice["K1"].value = "MOP"
    invoice["L1"].value = "Profit"
    invoice["M1"].value = "Bill No."
    invoice["N1"].value = "Date"
    invoice["O1"].value = "Time"

    art["A1"].value = "Art no."
    art["B1"].value = "Art Desc."
    art["C1"].value = "MRP"
    art["D1"].value = "OUM"
    art["E1"].value = "CP"

    wb.save("DATABASE.xlsx")
    os.startfile("DATABASE.xlsx")
    exit()
# Enter E --> open excel
clear()
inp1 = input("Enter E to open excel and Enter key to continue\t")
if inp1 == "E" or inp1 == "e":
    os.startfile("DATABASE.xlsx")
    clear()
    os.system("main.py")
    exit()
else:
    # loads details if customer exists
    X = 1
    phno = str(input("input Ph no {format - 91+XXXXXXXXXX}\t"))
    for row in range(1, cust.max_row + 1):
        pch = cust["A"+str(row)].value
        if pch == phno:
            row_exists = row  # row in which contact exists
            name = cust["B"+str(row)].value
            bday = cust["C"+str(row)].value
            bmon = cust["D"+str(row)].value
            aday = cust["F"+str(row)].value
            amon = cust["G"+str(row)].value
            frequency = cust["I"+str(row)].value + 1
            exists = True
            break
        if pch != phno and row == cust.max_row:
            name = str(input("Enter name \t"))
            bday = int(input("Enter birth date\t"))
            bmon = int(input("Enter birth month\t"))
            aday = int(input("Enter anniversary date\t"))
            amon = int(input("Enter anniversary month\t"))
            frequency = 1
            exists = False
            break
    if bill.max_row == 1:
        billno = 1
    elif bill.max_row != 1:
        billno = bill.max_row  # bill number for current bill
        print(billno)
    X = datetime.now()
    d = int(X.strftime("%d"))
    m = int(X.strftime("%m"))
    H = int(X.strftime("%H"))
    M = int(X.strftime("%M"))
    date = datetime.today()
    time_now = X.strftime("%H:%M")
    MOP = str(input("method of payment\t"))
    noA = int(input("No. of articles\t"))
    Y = 0  # count if Noa reaches --> turn while loop off
    total_shopping = 0  # total SP collected over time
    prev_profit = 0
    total_CP = 0  # total CP costed over time
    # get total sp and cp
    for Row in range(1, bill.max_row + 1):
        pch = bill["A" + str(Row)].value
        if pch == phno:
            total_shopping += bill["F" + str(Row)].value  # total SP
            prev_profit = cust["J"+str(row_exists)].value  # prev. earned profit from the cust.
            total_CP += bill["H" + str(Row)].value  # total cp
    X = 1
    SP_all = 0
    CP_all = 0
    while Y != noA:
        Art_no = int(input("Enter the same article number\t"))
        ran = 0  # for getting how many rows filled / ran this X no. of times
        for ROW in range(1, art.max_row + 1):
            ART_NO = art["A" + str(ROW)].value
            # to check if article number is valid
            if Art_no == ART_NO:
                desc = art["B" + str(ROW)].value
                MRP = art["C" + str(ROW)].value
                OUM = art["D" + str(ROW)].value
                CP = art["E"+str(ROW)].value
                break
            elif Art_no != ART_NO and ROW == art.max_row:
                print("ENTER A LEGITIMATE ARTICLE NUMBER NEXT TIME")
                time.sleep(2)
                clear()
                os.system("main.py")
                exit()
        QNT = int(input("Enter no. of similar article\t"))
        amount = QNT * MRP  # cost of same articles
        constant = (total_shopping / (frequency*300))  # constant for getting discount --> CHANGE
        discount = frequency + constant  # discount determination
        if discount >= 45:
            discount = 45
        if d == bday or d == aday:
            if m == bmon or m == amon:
                discount = 20 + discount  # special occasion discount
        SP = amount - amount * discount / 100  # SP + CP != MRP , SP can be more... for profit
        CP = CP * QNT
        SP_all += SP
        CP_all += CP
        profit_all = SP_all - CP_all
        Y = Y + QNT  # no. of ART billed
        print("Cost of " + str(QNT) + " " + str(OUM) + " " + str(desc) + " is " + str(SP))
        if Y == noA:
            profit_now = SP_all - CP_all  # profit earned now --> wrong
            while X == 1:
                if exists == False:
                    profit_cust = prev_profit + profit_all  # profit over time inc. now
                    cust.append([phno, name, bday, bmon, "U", aday, amon, "U", frequency, profit_cust])
                    X = 0
                elif exists == True:
                    cust["J"+str(row_exists)].value = profit_now + prev_profit
                    cust["I"+str(row_exists)].value = frequency
                    X = 0
        SP = int(SP)
        CP = int(CP)
        profit = SP - CP
        profit = int(profit)
        invoice.append([phno, Art_no, desc, QNT, OUM, MRP, amount, discount, SP, CP, MOP, profit, billno, date, time_now])
        ran = ran + 1
    bill.append([phno, d, m, H, M, SP_all, MOP, CP_all, profit_now, billno])  # has to be out to enter once
    price = 0
    print(invoice.max_row)
    print("TOTAL COST = " + str(SP_all))
    wb.save('DATABASE.xlsx')
    print("SAVED")
    db = pd.read_excel("DATABASE.xlsx", "Invoice", dtype=str, index_col=None)
    # db.to_csv('Machine Learning Database.csv', encode=utf_8_encode, header=True)
    print("\n\n\n ***THANKS FOR SHOPPING WITH US***\n\n\n")
    time.sleep(10)
    clear()
    os.system("main.py")
    os.system('background.py')
    exit()
