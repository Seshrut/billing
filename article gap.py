# it generares an excel sheet which contains art. no. and avr gap of it sales by each cust.

# import modules
import datetime
from datetime import *
import os
import openpyxl
from openpyxl import *
import pandas as pd
from pandas import *

# load workbook and worksheet
workbook = load_workbook('DATABASE.xlsx')
invoice = workbook['Invoice']
art = workbook['Articles']
cust = workbook['Customer']

a = 1
'''
Step 1 => chk_art = 1
Step 2 => look in invoice for art no chk_art
Step 3 => copy date, phno and art num of all rows having the chk_art
step 4 => paste in a temp file
step 5 => arrange in order of phno and chk the dates 
Step 6 => check the gap in dates for a perticular phno.
Step 7 => save that gap's mean in another file's row - | phno | art.no | gap mean |
Step 8 => chk_art = chk_art + 1
Step 9 => if chk_art = articles.max()
Step 10=> then exit()
Step 11 => Goto step 1... set chk_art = 2
'''

# step 1 - step 4
X = 0
for chk_art in range(0,int(art["A"+str(art.max_row)].value)):  # total articles
    WB = openpyxl.Workbook()
    WS = WB.active
    for row_inv in range(2,int(invoice.max_row)+1):  # rows in invoice
        
        if invoice["B"+str(row_inv)].value == chk_art:
            phno = str(invoice["A"+str(row_inv)].value)
            Date = invoice["N"+str(row_inv)].value

            WS["A1"].value = "phno"
            WS["B1"].value = "Date"
            WS["C1"].value = "Article no."

            WS.append([phno, Date, chk_art])

        if row_inv == invoice.max_row and chk_art != invoice["B"+str(row_inv)].value and WS.max_row > 1:
            for column_format in range (1,(WS.max_row)+1):
                column_format = WS['B'+str(column_format)]
                column_format.number_format = 'general'
            WB.save('temp '+str(chk_art)+'.xlsx')
# step 5
            try:
                Wb = pd.read_excel('temp '+str(chk_art)+'.xlsx') #sorting
                Wb = Wb.sort_values(by=['Date'])
                Wb.to_excel('temp '+str(chk_art)+'.xlsx',sheet_name='Sheet1')
                print(Wb)
                print('sorted ','temp '+str(chk_art)+'.xlsx')
            except Exception as E:
                print(E)
# step 6
# Find out mean for pert phno and save in a file with art no.
for FileName in range(2,art.max_row+1):
    fileName = FileName
    FileName = "temp " + str(art['A'+str(FileName)].value) + ".xlsx"
    try:
        WB = load_workbook(FileName)
        WS = WB.active
        for phno in range(2,WS.max_row+1):
            row_phno = phno
            date = WS['B'+str(phno)].value
            phno = WS['A'+str(phno)].value
            for phno_chk in range(2,WS.max_row+1):
                row_phno_chk = phno_chk
                date_chk = WS['B'+str(row_phno_chk)].value
                phno_chk = WS['A'+str(row_phno_chk)].value
                if phno == phno_chk and row_phno_chk > row_phno and row_phno != row_phno_chk:
                    dategap = date_chk - date
                    dategap = str(dategap)
                    for letter in range(1,int(len(dategap)+1)):  # runs only once
                        print(letter)
                        com = dategap[:letter]
                        if com[letter] == ",":
                            print(com)
                            com_len = len(com)
                            dategap = com[:com_len-1]
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.append([phno,dategap,fileName])
                    wb.save('MEAN.xlsx')
                    print(dategap)
    except Exception as E:
        b = 1

a = input()
