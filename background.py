import datetime
import pyautogui
import pywhatkit
import os
from openpyxl import load_workbook


def clear():
    print("\n"*30)


X = datetime.datetime.now()
d = int(X.strftime("%d"))  # date
m = int(X.strftime("%m"))  # month
H = int(X.strftime("%H"))  # hour
M = int(X.strftime("%M"))  # minute

wb = load_workbook("DATABASE.xlsx")
ws = wb['Customers']
disc = 45

for row in range(1, ws.max_row+1):
    B_day = ws["C"+str(row)].value
    B_mon = ws["D"+str(row)].value
    A_day = ws["F"+str(row)].value
    A_mon = ws["G"+str(row)].value
    if B_day == d and B_mon == m and ws["E"+str(row)].value == "U":
        ph_no = ws["A"+str(row)].value
        name = ws["B"+str(row)].value
        loyalty = ws["I"+str(row)].value
        discount = loyalty + 20
        if discount > 45:
            discount = 45
        message = str("Happy birthday " + str(name) + " May all your wishes come true. Today at this special occasion, We provide you with " + discount + "% OFF on everything SHOP AT ***** ENTERPRISE")
        G = pyautogui.position()
        pywhatkit.sendwhatmsg(ph_no, message, 00, M + 2, 30)
        pyautogui.moveTo(1063, 700)
        pyautogui.click(1063, 700)
        pyautogui.press('enter')
        pyautogui.moveTo(G)
        ws["E"+str(row)].vlaue = "S"
        wb.save("DATABASE.xlsx")
    if A_day == d and B_mon == m and ws["H"+str(row)].value == "U":
        ph_no = ws["A"+str(row)].value
        name = ws["B"+str(row)].value
        loyalty = ws["I"+str(row)].value
        discount = loyalty + 20
        if discount > 45:
            discount = 45
        message = str("Happy anniversary " + name + " May all your wishes come true. Today at this special occasion, We provide you and your loved one with " + discount + "% OFF on everything SHOP AT ***** ENTERPRISE")
        G = pyautogui.position()
        pywhatkit.sendwhatmsg(ph_no, message, 00, M + 2, 30)
        pyautogui.moveTo(1063, 700)
        pyautogui.click(1063, 700)
        pyautogui.press('enter')
        pyautogui.moveTo(G)
        ws["H"+str(row)].vlaue = "S"
        wb.save("DATABASE")
    if row == ws.max_row:
        break
if d == 1 and H == 0 and M == 1:
    for row in range(2, ws.max_row+1):
        ws["I"+str(row)].value = 0
if H == 0 and M == 0:
    for row in range(2, ws.max_row+1):
        ws["E"+str(row)].value = "U"
        ws["H"+str(row)].value = "U"
        break
print("I AM DONE")
clear()
os.system("main.py")
exit()
